"""
Dog Head Orientation - Manual Labeling Tool (v2)
=================================================

A robust video-based UI for manually reviewing and correcting dog head orientation labels.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import cv2
from PIL import Image, ImageTk
import pandas as pd
import numpy as np
from pathlib import Path
import math


class ManualLabelingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Dog Head Orientation - Manual Labeler")
        self.root.geometry("1200x750")
        
        # Project directories
        self.project_dir = Path(__file__).parent
        self.data_dir = self.project_dir / "data"
        self.output_dir = self.project_dir / "output"
        self.data_dir.mkdir(exist_ok=True)
        self.output_dir.mkdir(exist_ok=True)
        
        # State variables
        self.video_path = None
        self.excel_path = None
        self.cap = None
        self.total_frames = 0
        self.fps = 30.0
        self.video_duration = 0
        self.current_index = 0  # Current time segment index
        self.frame_interval = 1.0  # Default interval in seconds (1s, 0.5s, or 0.2s)
        self.total_segments = 0  # Total number of segments based on interval
        
        # Data
        self.df = None
        self.labels = {}  # index -> orientation
        self.auto_labels = {}  # auto-predicted labels
        self.angle_data = {}  # index -> (angle, is_straight)
        self.nose_landmarks = {}  # index -> dict of nose landmark positions
        self.border_data = {}  # index -> (x_bordo, y_bordo) from ray intersection
        
        # UI state
        self.photo = None  # Keep reference to prevent garbage collection
        
        self.create_ui()
        
    def create_ui(self):
        """Create the main UI layout."""
        # Configure grid weights
        self.root.columnconfigure(0, weight=3)
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(1, weight=1)
        
        # === TOP BAR - File Selection ===
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.grid(row=0, column=0, columnspan=2, sticky="ew")
        
        ttk.Button(top_frame, text="📂 Load Video", command=self.load_video).pack(side="left", padx=5)
        ttk.Button(top_frame, text="📊 Load Excel", command=self.load_excel).pack(side="left", padx=5)
        
        # Frame interval selector
        ttk.Label(top_frame, text="Interval:").pack(side="left", padx=(15, 2))
        self.interval_var = tk.StringVar(value="1.0s")
        interval_combo = ttk.Combobox(top_frame, textvariable=self.interval_var, 
                                       values=["1.0s", "0.5s", "0.2s"], width=6, state="readonly")
        interval_combo.pack(side="left", padx=2)
        interval_combo.bind("<<ComboboxSelected>>", self.on_interval_change)
        
        ttk.Button(top_frame, text="▶ Start Labeling", command=self.start_labeling).pack(side="left", padx=20)
        
        self.status_var = tk.StringVar(value="Load video and excel files to begin")
        ttk.Label(top_frame, textvariable=self.status_var, foreground="gray").pack(side="left", padx=20)
        
        # === LEFT SIDE - Video Display ===
        video_frame = ttk.LabelFrame(self.root, text="Video Preview", padding="10")
        video_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        video_frame.columnconfigure(0, weight=1)
        video_frame.rowconfigure(0, weight=1)
        
        # Canvas for video
        self.canvas = tk.Canvas(video_frame, bg="black", width=700, height=500)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        
        # Navigation controls
        nav_frame = ttk.Frame(video_frame)
        nav_frame.grid(row=1, column=0, pady=10)
        
        ttk.Button(nav_frame, text="⏮ First", command=self.go_first, width=8).pack(side="left", padx=2)
        ttk.Button(nav_frame, text="◀ Prev", command=self.go_prev, width=8).pack(side="left", padx=2)
        
        self.second_var = tk.StringVar(value="0 / 0")
        ttk.Label(nav_frame, textvariable=self.second_var, font=('Helvetica', 14, 'bold'), width=12).pack(side="left", padx=20)
        
        ttk.Button(nav_frame, text="Next ▶", command=self.go_next, width=8).pack(side="left", padx=2)
        ttk.Button(nav_frame, text="Last ⏭", command=self.go_last, width=8).pack(side="left", padx=2)
        
        # === RIGHT SIDE - Labeling Panel (Scrollable) ===
        right_outer_frame = ttk.Frame(self.root)
        right_outer_frame.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        
        # Create canvas with scrollbar for right panel
        right_canvas = tk.Canvas(right_outer_frame, width=220, highlightthickness=0)
        right_scrollbar = ttk.Scrollbar(right_outer_frame, orient="vertical", command=right_canvas.yview)
        right_frame = ttk.Frame(right_canvas, padding="5")
        
        # Configure scrolling
        right_frame.bind("<Configure>", lambda e: right_canvas.configure(scrollregion=right_canvas.bbox("all")))
        right_canvas.create_window((0, 0), window=right_frame, anchor="nw")
        right_canvas.configure(yscrollcommand=right_scrollbar.set)
        
        # Pack scrollbar and canvas
        right_scrollbar.pack(side="right", fill="y")
        right_canvas.pack(side="left", fill="both", expand=True)
        
        # Enable mousewheel scrolling
        def on_mousewheel(event):
            right_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        right_canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Current Second Display (compact)
        sec_frame = ttk.Frame(right_frame)
        sec_frame.pack(pady=(0, 5))
        ttk.Label(sec_frame, text="Second:", font=('Helvetica', 10)).pack(side="left")
        self.big_second_var = tk.StringVar(value="-")
        ttk.Label(sec_frame, textvariable=self.big_second_var, font=('Helvetica', 24, 'bold')).pack(side="left", padx=5)
        
        # Auto-detected label (compact)
        auto_frame = ttk.Frame(right_frame)
        auto_frame.pack(pady=2)
        ttk.Label(auto_frame, text="Auto:", font=('Helvetica', 9)).pack(side="left")
        self.auto_label_var = tk.StringVar(value="-")
        ttk.Label(auto_frame, textvariable=self.auto_label_var, font=('Helvetica', 12, 'bold'), foreground="blue").pack(side="left", padx=5)
        
        # Separator
        ttk.Separator(right_frame, orient="horizontal").pack(fill="x", pady=5)
        
        # Label Selection - 2x2 grid for compactness
        ttk.Label(right_frame, text="Select Label:", font=('Helvetica', 10, 'bold')).pack(pady=(0, 5))
        
        buttons_frame = ttk.Frame(right_frame)
        buttons_frame.pack(pady=5)
        
        self.label_buttons = {}
        button_configs = [
            ("LEFT", "#FF6B6B", "1", 0, 0),
            ("RIGHT", "#4ECDC4", "2", 0, 1),
            ("STRAIGHT", "#95E1A3", "3", 1, 0),
            ("ELSEWHERE", "#B0B0B0", "4", 1, 1),
        ]
        
        for label_name, color, key, row, col in button_configs:
            btn = tk.Button(
                buttons_frame, 
                text=f"{label_name}\n({key})",
                bg=color,
                font=('Helvetica', 9, 'bold'),
                width=10,
                height=2,
                command=lambda l=label_name: self.set_current_label(l)
            )
            btn.grid(row=row, column=col, padx=3, pady=3)
            self.label_buttons[label_name] = btn
        
        # Confirm & Next button
        ttk.Button(right_frame, text="✓ Confirm & Next (Enter)", 
                   command=self.confirm_and_next).pack(pady=10)
        
        # Separator
        ttk.Separator(right_frame, orient="horizontal").pack(fill="x", pady=5)
        
        # Progress (compact)
        prog_frame = ttk.Frame(right_frame)
        prog_frame.pack(pady=5)
        ttk.Label(prog_frame, text="Progress:", font=('Helvetica', 9)).pack(side="left")
        self.progress_var = tk.StringVar(value="0 / 0")
        ttk.Label(prog_frame, textvariable=self.progress_var, font=('Helvetica', 9, 'bold')).pack(side="left", padx=5)
        
        self.progress_bar = ttk.Progressbar(right_frame, length=180, mode='determinate')
        self.progress_bar.pack(pady=3)
        
        # Separator
        ttk.Separator(right_frame, orient="horizontal").pack(fill="x", pady=5)
        
        # === HEAD ORIENTATION INDICATOR ===
        angle_frame = ttk.LabelFrame(right_frame, text="Head Orientation (Nose Landmarks)", padding="3")
        angle_frame.pack(pady=3, fill="x")
        
        # Canvas for nose landmarks visualization
        self.angle_canvas = tk.Canvas(angle_frame, width=180, height=100, bg="#1a1a2e")
        self.angle_canvas.pack(pady=2)
        
        # Orientation info display
        info_frame = ttk.Frame(angle_frame)
        info_frame.pack(pady=2, fill="x")
        
        # X difference display
        self.y_diff_var = tk.StringVar(value="ΔX: --")
        ttk.Label(info_frame, textvariable=self.y_diff_var, font=('Helvetica', 8)).pack(side="left", padx=5)
        
        # Detected orientation
        self.detected_orient_var = tk.StringVar(value="--")
        ttk.Label(info_frame, textvariable=self.detected_orient_var, font=('Helvetica', 9, 'bold'), foreground="purple").pack(side="right", padx=5)
        
        self.tilt_status_var = tk.StringVar(value="")
        ttk.Label(angle_frame, textvariable=self.tilt_status_var, font=('Helvetica', 7), foreground="gray").pack()
        
        # === HEAD TILT ANGLE INDICATOR ===
        tilt_frame = ttk.LabelFrame(right_frame, text="Head Tilt Angle", padding="3")
        tilt_frame.pack(pady=3, fill="x")
        
        # Canvas for tilt angle visualization (larger to show full angle)
        self.tilt_canvas = tk.Canvas(tilt_frame, width=180, height=120, bg="white")
        self.tilt_canvas.pack(pady=2)
        
        # Tilt angle display - more prominent
        tilt_info_frame = ttk.Frame(tilt_frame)
        tilt_info_frame.pack(pady=2, fill="x")
        
        self.angle_var = tk.StringVar(value="Angle: --")
        ttk.Label(tilt_info_frame, textvariable=self.angle_var, font=('Helvetica', 11, 'bold'), foreground="#4CAF50").pack(side="left", padx=5)
        
        self.tilt_direction_var = tk.StringVar(value="")
        ttk.Label(tilt_info_frame, textvariable=self.tilt_direction_var, font=('Helvetica', 9)).pack(side="right", padx=5)
        
        # === IMAGE SPACE VISUALIZATION ===
        image_frame = ttk.LabelFrame(right_frame, text="Image Space & Ray", padding="3")
        image_frame.pack(pady=3, fill="x")
        
        # Canvas for image space visualization
        self.image_space_canvas = tk.Canvas(image_frame, width=180, height=130, bg="white", highlightthickness=1, highlightbackground="gray")
        self.image_space_canvas.pack(pady=2)
        
        # Info text below canvas
        self.ray_hit_var = tk.StringVar(value="Ray hit: --")
        ttk.Label(image_frame, textvariable=self.ray_hit_var, font=('Helvetica', 8)).pack(pady=2)
        
        # SAVE BUTTON - prominent at bottom
        save_btn = tk.Button(
            right_frame, 
            text="💾 SAVE LABELS", 
            command=self.save_labels,
            bg="#4CAF50",
            fg="white",
            font=('Helvetica', 11, 'bold'),
            width=16,
            height=2
        )
        save_btn.pack(pady=10)
        
        # === BOTTOM - Labels Overview ===
        bottom_frame = ttk.LabelFrame(self.root, text="All Labels Overview", padding="5")
        bottom_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 10))
        
        # Scrollable frame for label buttons
        canvas_scroll = tk.Canvas(bottom_frame, height=60)
        scrollbar = ttk.Scrollbar(bottom_frame, orient="horizontal", command=canvas_scroll.xview)
        self.labels_frame = ttk.Frame(canvas_scroll)
        
        canvas_scroll.configure(xscrollcommand=scrollbar.set)
        scrollbar.pack(side="bottom", fill="x")
        canvas_scroll.pack(side="top", fill="x")
        
        self.labels_window = canvas_scroll.create_window((0, 0), window=self.labels_frame, anchor="nw")
        self.labels_canvas = canvas_scroll
        
        self.labels_frame.bind("<Configure>", lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all")))
        
        # Keyboard bindings
        self.root.bind('<Left>', lambda e: self.go_prev())
        self.root.bind('<Right>', lambda e: self.go_next())
        self.root.bind('<Return>', lambda e: self.confirm_and_next())
        self.root.bind('<Key-1>', lambda e: self.set_current_label("LEFT"))
        self.root.bind('<Key-2>', lambda e: self.set_current_label("RIGHT"))
        self.root.bind('<Key-3>', lambda e: self.set_current_label("STRAIGHT"))
        self.root.bind('<Key-4>', lambda e: self.set_current_label("ELSEWHERE"))
        
    def load_video(self):
        """Load video file."""
        path = filedialog.askopenfilename(
            title="Select Video",
            initialdir=str(self.data_dir),
            filetypes=[("Video files", "*.mp4 *.avi *.mov"), ("All files", "*.*")]
        )
        if path:
            self.video_path = path
            self.status_var.set(f"Video: {Path(path).name}")
            
    def load_excel(self):
        """Load Excel data file."""
        path = filedialog.askopenfilename(
            title="Select Excel Data",
            initialdir=str(self.data_dir),
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.excel_path = path
            self.status_var.set(f"Excel: {Path(path).name}")
            
    def on_interval_change(self, event=None):
        """Handle frame interval selection change."""
        interval_str = self.interval_var.get()
        self.frame_interval = float(interval_str.replace("s", ""))
        
        # If already labeling, restart with new interval
        if self.cap is not None and self.df is not None:
            self.load_and_process_excel()
            self.labels = self.auto_labels.copy()
            self.create_overview_buttons()
            self.current_index = 0
            self.show_segment(0)
            self.status_var.set(f"Ready! {self.total_segments} segments ({self.frame_interval}s each)")
    
    def start_labeling(self):
        """Initialize labeling after files are loaded."""
        if not self.video_path:
            messagebox.showerror("Error", "Please load a video file first")
            return
        if not self.excel_path:
            messagebox.showerror("Error", "Please load an Excel data file first")
            return
            
        try:
            # Open video
            self.cap = cv2.VideoCapture(self.video_path)
            if not self.cap.isOpened():
                raise Exception("Could not open video file")
                
            self.total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.fps = self.cap.get(cv2.CAP_PROP_FPS) or 30.0
            self.video_duration = self.total_frames / self.fps
            
            # Parse selected interval
            interval_str = self.interval_var.get()
            self.frame_interval = float(interval_str.replace("s", ""))
            
            # Load Excel and calculate auto labels
            self.load_and_process_excel()
            
            # Initialize labels with auto-detected values
            self.labels = self.auto_labels.copy()
            
            # Create overview buttons
            self.create_overview_buttons()
            
            # Show first segment
            self.current_index = 0
            self.show_segment(0)
            
            self.status_var.set(f"Ready! {self.total_segments} segments ({self.frame_interval}s each)")
            messagebox.showinfo("Ready", f"Loaded {self.total_segments} segments ({self.frame_interval}s each).\nUse 1-4 keys to label, Enter to confirm and advance.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to start: {str(e)}")
            
    def load_and_process_excel(self):
        """Load Excel data and calculate auto labels."""
        df = pd.read_excel(self.excel_path, header=None, skiprows=2)
        df.columns = [
            'frame', 
            'nose_tip_x', 'nose_tip_y', 'nose_tip_likelihood',
            'nose_right_x', 'nose_right_y', 'nose_right_likelihood',
            'nose_bottom_x', 'nose_bottom_y', 'nose_bottom_likelihood',
            'nose_left_x', 'nose_left_y', 'nose_left_likelihood'
        ]
        
        # Convert all coordinate columns to numeric (IMPORTANT!)
        coord_cols = [
            'nose_tip_x', 'nose_tip_y', 'nose_tip_likelihood',
            'nose_right_x', 'nose_right_y', 'nose_right_likelihood',
            'nose_bottom_x', 'nose_bottom_y', 'nose_bottom_likelihood',
            'nose_left_x', 'nose_left_y', 'nose_left_likelihood'
        ]
        
        for col in coord_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        self.df = df
        
        # Calculate total segments based on interval
        self.total_segments = max(1, int(self.video_duration / self.frame_interval))
        
        # Calculate frames per segment
        frames_per_segment = len(df) / self.total_segments
        
        # Clear previous data
        self.auto_labels = {}
        self.angle_data = {}
        self.nose_landmarks = {}
        self.border_data = {}
        
        for seg_idx in range(self.total_segments):
            start_idx = int(seg_idx * frames_per_segment)
            end_idx = int((seg_idx + 1) * frames_per_segment)
            segment_data = df.iloc[start_idx:end_idx]
            
            if len(segment_data) == 0:
                self.auto_labels[seg_idx] = "ELSEWHERE"
                self.nose_landmarks[seg_idx] = None
                self.angle_data[seg_idx] = (0, True)
                self.border_data[seg_idx] = (None, None)
                continue
            
            # Get average nose landmark positions for this segment
            avg_tip_x = segment_data['nose_tip_x'].mean()
            avg_tip_y = segment_data['nose_tip_y'].mean()
            avg_tip_likelihood = segment_data['nose_tip_likelihood'].mean()
            avg_right_x = segment_data['nose_right_x'].mean()
            avg_right_y = segment_data['nose_right_y'].mean()
            avg_left_x = segment_data['nose_left_x'].mean()
            avg_left_y = segment_data['nose_left_y'].mean()
            avg_bottom_x = segment_data['nose_bottom_x'].mean()
            avg_bottom_y = segment_data['nose_bottom_y'].mean()
            avg_bottom_likelihood = segment_data['nose_bottom_likelihood'].mean()
            
            # Store landmarks for visualization
            self.nose_landmarks[seg_idx] = {
                'tip': (avg_tip_x, avg_tip_y),
                'right': (avg_right_x, avg_right_y),
                'left': (avg_left_x, avg_left_y),
                'bottom': (avg_bottom_x, avg_bottom_y),
                'tip_likelihood': avg_tip_likelihood,
                'bottom_likelihood': avg_bottom_likelihood
            }
            
            # Classify using ray intersection with image border (based on X,Y coordinates only)
            x_border, y_border, orientation = self.classify_orientation_by_ray_intersection(
                avg_bottom_x, avg_bottom_y, avg_tip_x, avg_tip_y,
                bottom_likelihood=avg_bottom_likelihood,
                tip_likelihood=avg_tip_likelihood
            )
            
            # Store x_bordo and y_bordo values
            self.border_data[seg_idx] = (x_border, y_border)
            
            # If undefined or poor likelihood, mark as "poor likelihood"
            if orientation in ("undefined", "poor_likelihood"):
                orientation = "poor likelihood"
            
            # Capitalize for consistency
            orientation = orientation.upper()
            
            self.auto_labels[seg_idx] = orientation
            
            # Calculate head tilt angle
            angle, is_straight = self.calculate_head_tilt_angle(
                avg_tip_x, avg_tip_y, avg_bottom_x, avg_bottom_y
            )
            self.angle_data[seg_idx] = (angle, is_straight)
    
    def calculate_head_tilt_angle(self, tip_x, tip_y, bottom_x, bottom_y, margin=2.0):
        """
        Calculate the head tilt angle based on nose tip and bottom positions.
        
        If nose.bottom X ≈ nose.tip X (within margin), head is straight (0°).
        Otherwise, calculate the angle from vertical.

        Returns: (angle_degrees, is_straight)
        """
        # Calculate horizontal difference
        dx = tip_x - bottom_x
        dy = bottom_y - tip_y  # Note: y increases downward in image coordinates
        
        # Check if head is straight (within margin)
        if abs(dx) <= margin:
            return 0.0, True
        
        # Calculate angle from vertical
        # When head is straight, tip is directly above bottom (dx=0)
        # Angle is measured from the vertical line
        if dy != 0:
            angle_rad = math.atan(dx / dy)
            angle_deg = math.degrees(angle_rad)
        else:
            # Edge case: dy is 0, head is horizontal
            angle_deg = 90.0 if dx > 0 else -90.0
        
        return round(angle_deg, 1), False
    
    def classify_orientation_by_ray_intersection(self, bottom_x, bottom_y, tip_x, tip_y, 
                                                 bottom_likelihood=None, tip_likelihood=None):
        """
        Classify head orientation by computing where the nose axis ray intersects the image border.
        
        POST-CALCULATION LIKELIHOOD CHECK (PRIORITY RULE):
        If bottom_likelihood < 0.6 OR tip_likelihood < 0.6, returns "poor_likelihood"
        instead of orientation. Orientation only output when both likelihoods >= 0.6.
        
        Classification ranges (EXPLICIT AND ROBUST):
        - LEFT: 125 < x < 325
        - STRAIGHT: 325 < x < 600
        - RIGHT: 600 < x < 800
        - ELSEWHERE: x < 125 or x > 800
        """
        W, H = 920, 518
        
        # Convert all inputs to float
        try:
            bottom_x = float(bottom_x)
            bottom_y = float(bottom_y)
            tip_x = float(tip_x)
            tip_y = float(tip_y)
        except (ValueError, TypeError):
            return None, None, "undefined"
        
        # Check for NaN values
        if any(np.isnan(v) for v in [bottom_x, bottom_y, tip_x, tip_y]):
            return None, None, "undefined"
        
        # POST-CALCULATION LIKELIHOOD CHECK (PRIORITY RULE)
        # If either likelihood is provided and < 0.6, return poor_likelihood
        if bottom_likelihood is not None and bottom_likelihood < 0.6:
            return None, None, "poor_likelihood"
        if tip_likelihood is not None and tip_likelihood < 0.6:
            return None, None, "poor_likelihood"
        
        # Direction vector
        dx = tip_x - bottom_x
        dy = tip_y - bottom_y
        x1, y1 = bottom_x, bottom_y
        
        # Find all valid border intersections
        intersections = []
        
        if dx != 0:
            # Left border
            t = (0 - x1) / dx
            if t >= 0:
                y = y1 + t * dy
                if 0 <= y <= H - 1:
                    intersections.append((t, 0, y))
        
            # Right border
            t = (W - 1 - x1) / dx
            if t >= 0:
                y = y1 + t * dy
                if 0 <= y <= H - 1:
                    intersections.append((t, W - 1, y))
        
        if dy != 0:
            # Top border
            t = (0 - y1) / dy
            if t >= 0:
                x = x1 + t * dx
                if 0 <= x <= W - 1:
                    intersections.append((t, x, 0))
        
            # Bottom border
            t = (H - 1 - y1) / dy
            if t >= 0:
                x = x1 + t * dx
                if 0 <= x <= W - 1:
                    intersections.append((t, x, H - 1))
        
        if not intersections:
            return None, None, "undefined"
        
        # Get first intersection (smallest t)
        intersections.sort(key=lambda x: x[0])
        t, x_border, y_border = intersections[0]
        
        # Round x_border to handle floating point precision issues
        x_border = round(x_border, 1)
        
        # Classify based on explicit range boundaries
        # IMPORTANT: Check ranges in order, with explicit conditions
        if x_border > 125 and x_border < 325:
            # LEFT panel
            orientation = "left"
        elif x_border > 600 and x_border < 800:
            # RIGHT panel
            orientation = "right"
        elif x_border > 325 and x_border < 600:
            # STRAIGHT zone (between panels)
            orientation = "straight"
        else:
            # ELSEWHERE (x_border < 125 or x_border > 800)
            orientation = "elsewhere"
        
        return x_border, y_border, orientation
    
    def update_angle_display(self):
        """Update the nose landmarks visualization and orientation info."""
        landmarks = self.nose_landmarks.get(self.current_index)
        angle_data = self.angle_data.get(self.current_index)
        
        if landmarks is None or angle_data is None:
            self.y_diff_var.set("Ray: --")
            self.detected_orient_var.set("No data")
            self.angle_var.set("Angle: --")
            self.tilt_status_var.set("")
            self.tilt_direction_var.set("")
            self.draw_empty_canvas()
            self.draw_empty_tilt_canvas()
            return
        
        # Get orientation from ray intersection
        orientation = self.auto_labels.get(self.current_index, "--")
        self.detected_orient_var.set(f"→ {orientation}")
        
        # Compute ray intersection for display
        tip = landmarks['tip']
        bottom = landmarks['bottom']
        tip_likelihood = landmarks.get('tip_likelihood', None)
        bottom_likelihood = landmarks.get('bottom_likelihood', None)
        
        # Calculate delta X (difference in x-coordinates between tip and bottom)
        delta_x = tip[0] - bottom[0]
        self.y_diff_var.set(f"ΔX: {delta_x:.1f}")
        
        x_border, y_border, _ = self.classify_orientation_by_ray_intersection(
            bottom[0], bottom[1], tip[0], tip[1],
            bottom_likelihood=bottom_likelihood,
            tip_likelihood=tip_likelihood
        )
        
        
        # Update tilt angle
        angle, is_straight = angle_data
        if is_straight:
            self.angle_var.set("Angle: 0°")
            self.tilt_direction_var.set("(Straight)")
        else:
            self.angle_var.set(f"Angle: {abs(angle):.1f}°")
            direction = "tilted LEFT" if angle < 0 else "tilted RIGHT"
            self.tilt_direction_var.set(f"({direction})")
        
        # Status text showing panel ranges
        if orientation == "LEFT":
            self.tilt_status_var.set("Left panel: (125, 325)")
        elif orientation == "RIGHT":
            self.tilt_status_var.set("Right panel: (600, 800)")
        elif orientation == "STRAIGHT":
            self.tilt_status_var.set("Between panels: (325, 600)")
        elif orientation == "POOR LIKELIHOOD":
            self.tilt_status_var.set("Likelihood < 0.6")
        else:
            self.tilt_status_var.set("Outside panels: <125 or >800")
        
        # Draw nose landmarks visualization
        self.draw_nose_landmarks(landmarks, orientation)
        
        # Draw tilt angle visualization
        self.draw_tilt_angle(landmarks, angle, is_straight)
        
        # Draw image space with ray visualization
        self.draw_image_space_with_ray(landmarks, orientation)
    
    def draw_empty_canvas(self):
        """Draw empty canvas with placeholder."""
        self.angle_canvas.delete("all")
        w = self.angle_canvas.winfo_width() or 180
        h = self.angle_canvas.winfo_height() or 80
        self.angle_canvas.create_text(w//2, h//2, text="No data", fill="gray", font=('Helvetica', 10))
    
    def draw_empty_tilt_canvas(self):
        """Draw empty tilt canvas with placeholder."""
        self.tilt_canvas.delete("all")
        w = self.tilt_canvas.winfo_width() or 180
        h = self.tilt_canvas.winfo_height() or 120
        self.tilt_canvas.create_text(w//2, h//2, text="No data", fill="gray", font=('Helvetica', 10))
    
    def draw_tilt_angle(self, landmarks, angle, is_straight):
        """Draw the head tilt angle visualization showing the triangle."""
        self.tilt_canvas.delete("all")
        
        w = self.tilt_canvas.winfo_width() or 180
        h = self.tilt_canvas.winfo_height() or 120
        
        # Get tip and bottom positions
        tip = landmarks['tip']
        bottom = landmarks['bottom']
        
        # Center point at bottom of visualization area
        cx = w // 2
        cy = h - 25
        line_length = 70  # Longer line for better visibility
        
        # Draw baseline (vertical reference line) - this is the "straight" reference
        baseline_top_x = cx
        baseline_top_y = cy - line_length
        
        self.tilt_canvas.create_line(
            cx, cy, baseline_top_x, baseline_top_y,
            fill="#AAAAAA", width=2, dash=(4, 2)
        )
        self.tilt_canvas.create_text(
            baseline_top_x - 12, baseline_top_y + 8,
            text="0°", fill="#888888", font=('Helvetica', 8)
        )
        
        # Calculate the actual head angle line
        angle_rad = math.radians(angle)
        actual_top_x = cx + line_length * math.sin(angle_rad)
        actual_top_y = cy - line_length * math.cos(angle_rad)
        
        # Color based on angle
        if is_straight:
            line_color = "#4CAF50"  # Green for straight
        elif abs(angle) < 10:
            line_color = "#FFC107"  # Yellow for slight tilt
        else:
            line_color = "#F44336"  # Red for significant tilt
        
        # Draw actual head angle line
        self.tilt_canvas.create_line(
            cx, cy, actual_top_x, actual_top_y,
            fill=line_color, width=3, arrow=tk.LAST
        )
        
        # Draw the triangle if tilted (show the angle)
        if not is_straight and abs(angle) > 1:
            # Draw horizontal line from baseline top to actual top (completes triangle)
            self.tilt_canvas.create_line(
                baseline_top_x, baseline_top_y, actual_top_x, actual_top_y,
                fill="#FFB6C1", width=1, dash=(2, 2)
            )
            
            # Draw angle arc
            arc_radius = 22
            start_angle = 90  # Vertical is 90 degrees in canvas coords
            extent = -angle  # Negative because canvas angles go counter-clockwise
            
            self.tilt_canvas.create_arc(
                cx - arc_radius, cy - arc_radius,
                cx + arc_radius, cy + arc_radius,
                start=start_angle, extent=extent,
                style=tk.ARC, outline=line_color, width=2
            )
        
        # Draw pivot point at bottom (larger)
        self.tilt_canvas.create_oval(
            cx - 5, cy - 5, cx + 5, cy + 5,
            fill=line_color, outline="white", width=1
        )
        
        # Labels for tip and bottom
        self.tilt_canvas.create_text(cx, cy + 12, text="bottom", fill="#666666", font=('Helvetica', 8))
        self.tilt_canvas.create_text(actual_top_x, actual_top_y - 10, text="tip", fill=line_color, font=('Helvetica', 8, 'bold'))
        
        # Draw prominent angle value in top-right corner
        angle_text = f"{abs(angle):.1f}°" if not is_straight else "0°"
        self.tilt_canvas.create_text(
            w - 10, 15,
            text=angle_text, fill=line_color, font=('Helvetica', 14, 'bold'),
            anchor="e"
        )
        
        # Draw tilt direction indicator
        if not is_straight:
            direction = "← LEFT" if angle < 0 else "RIGHT →"
            self.tilt_canvas.create_text(
                w - 10, 32,
                text=direction, fill=line_color, font=('Helvetica', 9),
                anchor="e"
            )

    def draw_nose_landmarks(self, landmarks, orientation):
        """Draw the nose landmarks visualization on the canvas."""
        self.angle_canvas.delete("all")
        
        w = self.angle_canvas.winfo_width() or 180
        h = self.angle_canvas.winfo_height() or 100
        
        # Get landmark positions
        tip = landmarks['tip']
        right = landmarks['right']
        left = landmarks['left']
        bottom = landmarks['bottom']
        
        # Find bounds of landmarks to scale them to canvas
        all_x = [tip[0], right[0], left[0], bottom[0]]
        all_y = [tip[1], right[1], left[1], bottom[1]]
        
        min_x, max_x = min(all_x), max(all_x)
        min_y, max_y = min(all_y), max(all_y)
        
        # Add padding
        padding = 15
        range_x = max_x - min_x if max_x != min_x else 1
        range_y = max_y - min_y if max_y != min_y else 1
        
        # Scale to fit canvas
        scale_x = (w - 2 * padding) / range_x
        scale_y = (h - 2 * padding) / range_y
        scale = min(scale_x, scale_y) * 0.8
        
        # Center offset
        cx = w // 2
        cy = h // 2
        center_x = (min_x + max_x) / 2
        center_y = (min_y + max_y) / 2
        
        def transform(x, y):
            """Transform real coordinates to canvas coordinates."""
            tx = cx + (x - center_x) * scale
            ty = cy + (y - center_y) * scale
            return tx, ty
        
        # Transform all points
        tip_t = transform(*tip)
        right_t = transform(*right)
        left_t = transform(*left)
        bottom_t = transform(*bottom)
        
        # Draw connecting lines (nose outline)
        line_color = "#555555"
        self.angle_canvas.create_line(*tip_t, *right_t, fill=line_color, width=1)
        self.angle_canvas.create_line(*tip_t, *left_t, fill=line_color, width=1)
        self.angle_canvas.create_line(*right_t, *bottom_t, fill=line_color, width=1)
        self.angle_canvas.create_line(*left_t, *bottom_t, fill=line_color, width=1)
        
        # Draw horizontal reference line between left and right (for Y comparison)
        # Color based on orientation
        orient_colors = {"LEFT": "#FF6B6B", "RIGHT": "#4ECDC4", "STRAIGHT": "#95E1A3", "ELSEWHERE": "#888888"}
        ref_color = orient_colors.get(orientation, "#888888")
        
        self.angle_canvas.create_line(
            left_t[0], left_t[1], right_t[0], right_t[1],
            fill=ref_color, width=3, dash=(5, 2)
        )
        
        # Draw Y level indicators
        # Left Y level line
        self.angle_canvas.create_line(
            left_t[0] - 15, left_t[1], left_t[0] - 5, left_t[1],
            fill="#FF6B6B", width=2
        )
        # Right Y level line  
        self.angle_canvas.create_line(
            right_t[0] + 5, right_t[1], right_t[0] + 15, right_t[1],
            fill="#4ECDC4", width=2
        )
        
        # Draw landmark points with labels
        point_radius = 5
        
        # Tip - Yellow
        self.angle_canvas.create_oval(
            tip_t[0] - point_radius, tip_t[1] - point_radius,
            tip_t[0] + point_radius, tip_t[1] + point_radius,
            fill="#FFD700", outline="white", width=1
        )
        self.angle_canvas.create_text(tip_t[0], tip_t[1] - 12, text="tip", fill="#FFD700", font=('Helvetica', 7))
        
        # Right - Cyan
        self.angle_canvas.create_oval(
            right_t[0] - point_radius, right_t[1] - point_radius,
            right_t[0] + point_radius, right_t[1] + point_radius,
            fill="#4ECDC4", outline="white", width=1
        )
        self.angle_canvas.create_text(right_t[0] + 12, right_t[1], text="R", fill="#4ECDC4", font=('Helvetica', 8, 'bold'))
        
        # Left - Red/Pink
        self.angle_canvas.create_oval(
            left_t[0] - point_radius, left_t[1] - point_radius,
            left_t[0] + point_radius, left_t[1] + point_radius,
            fill="#FF6B6B", outline="white", width=1
        )
        self.angle_canvas.create_text(left_t[0] - 12, left_t[1], text="L", fill="#FF6B6B", font=('Helvetica', 8, 'bold'))
        
        # Bottom - White
        self.angle_canvas.create_oval(
            bottom_t[0] - point_radius, bottom_t[1] - point_radius,
            bottom_t[0] + point_radius, bottom_t[1] + point_radius,
            fill="white", outline="#888888", width=1
        )
        self.angle_canvas.create_text(bottom_t[0], bottom_t[1] + 12, text="btm", fill="white", font=('Helvetica', 7))
        
        # Draw orientation indicator arrow at the top
        arrow_y = 10
        if orientation == "LEFT":
            self.angle_canvas.create_line(w//2, arrow_y, w//2 - 25, arrow_y, fill="#FF6B6B", width=3, arrow=tk.FIRST)
        elif orientation == "RIGHT":
            self.angle_canvas.create_line(w//2, arrow_y, w//2 + 25, arrow_y, fill="#4ECDC4", width=3, arrow=tk.LAST)
        elif orientation == "STRAIGHT":
            self.angle_canvas.create_oval(w//2 - 5, arrow_y - 5, w//2 + 5, arrow_y + 5, fill="#95E1A3", outline="#95E1A3")
            
    def draw_image_space_with_ray(self, landmarks, orientation):
        """
        Draw the image space visualization showing:
        - Image bounds (920x518)
        - Panel ranges (left, right)
        - Nose landmarks
        - Ray from nose bottom through tip to border
        """
        self.image_space_canvas.delete("all")
        
        w = self.image_space_canvas.winfo_width() or 180
        h = self.image_space_canvas.winfo_height() or 130
        
        # Image dimensions
        IMG_W, IMG_H = 920, 518
        
        # Exact panel ranges (updated)
        PANEL_LEFT_MIN = 125
        PANEL_LEFT_MAX = 325
        PANEL_RIGHT_MIN = 600
        PANEL_RIGHT_MAX = 800
        STRAIGHT_MIN = 325
        STRAIGHT_MAX = 600
        
        # Padding on canvas
        padding = 10
        canvas_w = w - 2 * padding
        canvas_h = h - 2 * padding
        
        # Scale factors
        scale_x = canvas_w / IMG_W
        scale_y = canvas_h / IMG_H
        
        def img_to_canvas(x, y):
            """Convert image coordinates to canvas coordinates."""
            cx = padding + x * scale_x
            cy = padding + y * scale_y
            return cx, cy
        
        # Draw background (light gray)
        self.image_space_canvas.create_rectangle(
            padding, padding, padding + canvas_w, padding + canvas_h,
            fill="#f5f5f5", outline="black", width=1
        )
        
        # Draw left panel range (red background)
        left_x1, _ = img_to_canvas(PANEL_LEFT_MIN, 0)
        left_x2, _ = img_to_canvas(PANEL_LEFT_MAX, 0)
        self.image_space_canvas.create_rectangle(
            left_x1, padding, left_x2, padding + canvas_h,
            fill="#FF6B6B", outline=""
        )
        self.image_space_canvas.create_text(
            (left_x1 + left_x2) / 2, padding + 5,
            text="L", fill="white", font=('Helvetica', 7, 'bold')
        )
        
        # Draw right panel range (cyan background)
        right_x1, _ = img_to_canvas(PANEL_RIGHT_MIN, 0)
        right_x2, _ = img_to_canvas(PANEL_RIGHT_MAX, 0)
        self.image_space_canvas.create_rectangle(
            right_x1, padding, right_x2, padding + canvas_h,
            fill="#4ECDC4", outline=""
        )
        self.image_space_canvas.create_text(
            (right_x1 + right_x2) / 2, padding + 5,
            text="R", fill="white", font=('Helvetica', 7, 'bold')
        )
        
        # Draw straight zone (green background - subtle)
        straight_x1, _ = img_to_canvas(STRAIGHT_MIN, 0)
        straight_x2, _ = img_to_canvas(STRAIGHT_MAX, 0)
        self.image_space_canvas.create_rectangle(
            straight_x1, padding, straight_x2, padding + canvas_h,
            fill="#95E1A3", outline=""
        )
        
        # Draw image border
        self.image_space_canvas.create_rectangle(
            padding, padding, padding + canvas_w, padding + canvas_h,
            outline="black", width=2
        )
        
        # Get landmarks
        tip = landmarks['tip']
        bottom = landmarks['bottom']
        
        # Draw nose landmarks
        tip_c = img_to_canvas(*tip)
        bottom_c = img_to_canvas(*bottom)
        
        # Draw line from bottom through tip and extend to border
        dx = tip[0] - bottom[0]
        dy = tip[1] - bottom[1]
        
        # Calculate ray to border
        if dx != 0 or dy != 0:
            # Find which border the ray hits
            intersections = []
            
            # Left border (x=0)
            if dx != 0:
                t = (0 - bottom[0]) / dx
                if t >= 0:
                    y = bottom[1] + t * dy
                    if 0 <= y <= IMG_H - 1:
                        intersections.append((t, 0, y))
            
            # Right border (x=919)
            if dx != 0:
                t = (IMG_W - 1 - bottom[0]) / dx
                if t >= 0:
                    y = bottom[1] + t * dy
                    if 0 <= y <= IMG_H - 1:
                        intersections.append((t, IMG_W - 1, y))
            
            # Top border (y=0)
            if dy != 0:
                t = (0 - bottom[1]) / dy
                if t >= 0:
                    x = bottom[0] + t * dx
                    if 0 <= x <= IMG_W - 1:
                        intersections.append((t, x, 0))
            
            # Bottom border (y=517)
            if dy != 0:
                t = (IMG_H - 1 - bottom[1]) / dy
                if t >= 0:
                    x = bottom[0] + t * dx
                    if 0 <= x <= IMG_W - 1:
                        intersections.append((t, x, IMG_H - 1))
            
            if intersections:
                # Get first intersection (smallest t)
                intersections.sort(key=lambda x: x[0])
                t, x_border, y_border = intersections[0]
                border_c = img_to_canvas(x_border, y_border)
                
                # Draw ray line from bottom through border
                self.image_space_canvas.create_line(
                    bottom_c[0], bottom_c[1], border_c[0], border_c[1],
                    fill="purple", width=2
                )
                
                # Draw marker at border intersection
                marker_r = 3
                self.image_space_canvas.create_oval(
                    border_c[0] - marker_r, border_c[1] - marker_r,
                    border_c[0] + marker_r, border_c[1] + marker_r,
                    fill="purple", outline="white", width=1
                )
                
                # Update ray hit information with likelihood check
                likelihood_note = ""
                if orientation == "elsewhere":
                    # Check if it's poor likelihood
                    tip_likelihood = landmarks.get('tip_likelihood', None)
                    bottom_likelihood = landmarks.get('bottom_likelihood', None)
                    if (bottom_likelihood is not None and bottom_likelihood < 0.6) or \
                       (tip_likelihood is not None and tip_likelihood < 0.6):
                        likelihood_note = " [poor likelihood]"
                
                self.ray_hit_var.set(f"Ray hit x={x_border:.0f} ({orientation}){likelihood_note}")
        
        # Draw nose landmarks as points
        # Tip
        self.image_space_canvas.create_oval(
            tip_c[0] - 2, tip_c[1] - 2, tip_c[0] + 2, tip_c[1] + 2,
            fill="#FFD700", outline="black", width=1
        )
        
        # Bottom
        self.image_space_canvas.create_oval(
            bottom_c[0] - 2, bottom_c[1] - 2, bottom_c[0] + 2, bottom_c[1] + 2,
            fill="white", outline="black", width=1
        )
        
        # Draw axis labels
        self.image_space_canvas.create_text(
            padding - 5, padding + canvas_h + 3,
            text="0", fill="gray", font=('Helvetica', 6)
        )
        self.image_space_canvas.create_text(
            padding + canvas_w + 3, padding + canvas_h + 3,
            text="920", fill="gray", font=('Helvetica', 6)
        )
            
    def create_overview_buttons(self):
        """Create the overview buttons at the bottom."""
        # Clear existing
        for widget in self.labels_frame.winfo_children():
            widget.destroy()
            
        self.overview_buttons = {}
        for seg_idx in range(self.total_segments):
            # Display time in seconds for button label
            time_sec = seg_idx * self.frame_interval
            if self.frame_interval == 1.0:
                label_text = str(seg_idx)
            else:
                label_text = f"{time_sec:.1f}"
            
            btn = tk.Button(
                self.labels_frame,
                text=label_text,
                width=4,
                height=2,
                font=('Helvetica', 7),
                command=lambda s=seg_idx: self.jump_to_segment(s)
            )
            btn.pack(side="left", padx=1)
            self.overview_buttons[seg_idx] = btn
            
        self.update_overview_colors()
        
    def update_overview_colors(self):
        """Update overview button colors based on labels."""
        colors = {
            "LEFT": "#FF6B6B",
            "RIGHT": "#4ECDC4", 
            "STRAIGHT": "#95E1A3",
            "ELSEWHERE": "#B0B0B0",
            "POOR LIKELIHOOD": "#FFB6C1"
        }
        
        for seg_idx, btn in self.overview_buttons.items():
            label = self.labels.get(seg_idx, "ELSEWHERE")
            btn.configure(bg=colors.get(label, "#FFFFFF"))
            
            # Highlight current
            if seg_idx == self.current_index:
                btn.configure(relief="sunken", borderwidth=3)
            else:
                btn.configure(relief="raised", borderwidth=1)
                
    def show_segment(self, seg_idx):
        """Display a specific segment of video."""
        if self.cap is None or seg_idx < 0 or seg_idx >= self.total_segments:
            return
            
        self.current_index = seg_idx
        
        # Calculate the time position for this segment (middle of segment)
        time_sec = (seg_idx + 0.5) * self.frame_interval
        frame_idx = int(time_sec * self.fps)
        frame_idx = min(frame_idx, self.total_frames - 1)
        
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        ret, frame = self.cap.read()
        
        if ret:
            # Convert to RGB
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            
            # Get canvas size
            self.canvas.update_idletasks()
            canvas_w = self.canvas.winfo_width()
            canvas_h = self.canvas.winfo_height()
            
            if canvas_w > 10 and canvas_h > 10:
                # Resize maintaining aspect ratio
                h, w = frame.shape[:2]
                scale = min(canvas_w / w, canvas_h / h) * 0.95
                new_w, new_h = int(w * scale), int(h * scale)
                frame = cv2.resize(frame, (new_w, new_h))
            
            # Add text overlay with time info
            time_start = seg_idx * self.frame_interval
            time_end = (seg_idx + 1) * self.frame_interval
            label = self.labels.get(seg_idx, "-")
            cv2.putText(frame, f"Time: {time_start:.1f}s - {time_end:.1f}s", (10, 35), 
                       cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 0), 2)
            cv2.putText(frame, f"Segment: {seg_idx} | Label: {label}", (10, 70),
                       cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
            
            # Display
            image = Image.fromarray(frame)
            self.photo = ImageTk.PhotoImage(image)
            
            self.canvas.delete("all")
            self.canvas.create_image(canvas_w//2, canvas_h//2, image=self.photo, anchor="center")
        
        # Update UI
        self.update_ui_state()
        
    def update_ui_state(self):
        """Update all UI elements for current state."""
        # Segment display with time info
        time_sec = self.current_index * self.frame_interval
        self.second_var.set(f"{self.current_index} / {self.total_segments - 1}")
        self.big_second_var.set(f"{time_sec:.1f}s")
        
        # Auto label
        auto = self.auto_labels.get(self.current_index, "-")
        self.auto_label_var.set(auto)
        
        # Highlight selected label button
        current_label = self.labels.get(self.current_index, "STRAIGHT")
        for name, btn in self.label_buttons.items():
            if name == current_label:
                btn.configure(relief="sunken", borderwidth=4)
            else:
                btn.configure(relief="raised", borderwidth=2)
        
        # Progress
        labeled = len(self.labels)
        total = self.total_segments
        self.progress_var.set(f"{labeled} / {total}")
        self.progress_bar['value'] = (labeled / total * 100) if total > 0 else 0
        
        # Overview
        self.update_overview_colors()
        
        # Update angle display
        self.update_angle_display()
        
    def set_current_label(self, label):
        """Set label for current segment."""
        self.labels[self.current_index] = label
        self.update_ui_state()
        self.show_segment(self.current_index)  # Refresh display
        
    def confirm_and_next(self):
        """Confirm current and go to next segment."""
        if self.current_index < self.total_segments - 1:
            self.go_next()
        else:
            messagebox.showinfo("Done", "You've reached the last segment!\nClick 'Save Labels' when ready.")
            
    def go_prev(self):
        """Go to previous segment."""
        if self.current_index > 0:
            self.show_segment(self.current_index - 1)
            
    def go_next(self):
        """Go to next segment."""
        if self.current_index < self.total_segments - 1:
            self.show_segment(self.current_index + 1)
            
    def go_first(self):
        """Go to first segment."""
        self.show_segment(0)
        
    def go_last(self):
        """Go to last segment."""
        self.show_segment(self.total_segments - 1)
        
    def jump_to_segment(self, seg_idx):
        """Jump to a specific segment."""
        self.show_segment(seg_idx)
        
    def save_labels(self):
        """Save labels to Excel with summary in upper rows."""
        if not self.labels:
            messagebox.showerror("Error", "No labels to save!")
            return
            
        # Generate filename
        input_name = Path(self.excel_path).stem if self.excel_path else "output"
        default_name = f"{input_name}_manual_labels.xlsx"
        
        path = filedialog.asksaveasfilename(
            title="Save Labels",
            initialdir=str(self.output_dir),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        
        if path:
            # Create DataFrame with segment data (including x_bordo and y_bordo)
            data = []
            for seg_idx in range(self.total_segments):
                time_start = seg_idx * self.frame_interval
                time_end = (seg_idx + 1) * self.frame_interval
                x_bordo, y_bordo = self.border_data.get(seg_idx, (None, None))
                data.append({
                    'segment': seg_idx,
                    'time_start': time_start,
                    'time_end': time_end,
                    'interval': self.frame_interval,
                    'orientation': self.labels.get(seg_idx, "ELSEWHERE"),
                    'x_bordo': x_bordo,
                    'y_bordo': y_bordo
                })
            
            df = pd.DataFrame(data)
            
            # Calculate percentages for each orientation
            total = len(df)
            orientation_counts = df['orientation'].value_counts()
            
            # Save to Excel with summary in upper rows starting from H2
            if path.endswith('.xlsx'):
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Labels"
                
                # Write headers
                headers = ['segment', 'time_start', 'time_end', 'interval', 'orientation', 'x_bordo', 'y_bordo']
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # Write data rows
                for row_idx, row_data in enumerate(data, 2):
                    ws.cell(row=row_idx, column=1, value=row_data['segment'])
                    ws.cell(row=row_idx, column=2, value=row_data['time_start'])
                    ws.cell(row=row_idx, column=3, value=row_data['time_end'])
                    ws.cell(row=row_idx, column=4, value=row_data['interval'])
                    ws.cell(row=row_idx, column=5, value=row_data['orientation'])
                    ws.cell(row=row_idx, column=6, value=row_data['x_bordo'])
                    ws.cell(row=row_idx, column=7, value=row_data['y_bordo'])
                
                # Write summary starting from J2 (moved to accommodate new columns)
                summary_row = 2
                ws.cell(row=1, column=10, value="SUMMARY").font = Font(bold=True)
                
                for orientation in ['LEFT', 'RIGHT', 'STRAIGHT', 'ELSEWHERE']:
                    count = orientation_counts.get(orientation, 0)
                    percentage = (count / total * 100) if total > 0 else 0
                    
                    ws.cell(row=summary_row, column=10, value=f"{orientation}")
                    ws.cell(row=summary_row, column=11, value=f"{count}")
                    ws.cell(row=summary_row, column=12, value=f"{percentage:.1f}%")
                    summary_row += 1
                
                # Save workbook
                wb.save(path)
            else:
                # Save as CSV if user selected .csv
                df.to_csv(path, index=False)
            
            # Show summary with percentages
            summary_msg = f"Labels saved to:\n{path}\n\n--- SUMMARY ---\n"
            for orientation in ['LEFT', 'RIGHT', 'STRAIGHT', 'ELSEWHERE']:
                count = orientation_counts.get(orientation, 0)
                percentage = (count / total * 100) if total > 0 else 0
                summary_msg += f"{orientation}: {count} ({percentage:.1f}%)\n"
            summary_msg += f"Total: {total}"
            
            messagebox.showinfo("Saved!", summary_msg)
            
    def on_close(self):
        """Cleanup on close."""
        if self.cap:
            self.cap.release()
        self.root.destroy()


def main():
    root = tk.Tk()
    app = ManualLabelingApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
